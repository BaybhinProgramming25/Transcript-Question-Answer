from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from database.database import get_db
from database.models import Document, Message
from helpers.jwt import get_current_user
from helpers.getchunks import parse_pdf
from helpers.limiter import limiter
from rag import init_db, delete_index
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import io
import os
import logging

OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
UPLOADS_DIR = os.getenv("UPLOADS_DIR", "/uploads")
router = APIRouter()
logger = logging.getLogger(__name__)


@router.post("/api/documents")
@limiter.limit("5/minute")
def upload_document(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):


    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")
    

    user_email = current_user["username"]
    try:
        if db.query(Document).filter(Document.user_email == user_email, Document.filename == file.filename).first():
            raise HTTPException(status_code=409, detail="A document with this name already exists")
    except SQLAlchemyError as e:
        logger.error(f"Could not query document with associated user: {e}")
        raise HTTPException(status_code=500, detail="Database error")
    

    filepath = None 
    try: 
        user_dir = os.path.join(UPLOADS_DIR, user_email)
        os.makedirs(user_dir, exist_ok=True)
        filepath = os.path.join(user_dir, os.path.basename(file.filename))
    except OSError as e:
        logger.error(f"Path Not found {e}")
        raise HTTPException(status_code=500, detail="Something went wrong")
    

    pdf_bytes = file.file.read()
    if not pdf_bytes.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")


    with open(filepath, "wb") as f:
        f.write(pdf_bytes)    
    size = len(pdf_bytes)


    doc = None 
    try: 
        doc = Document(
            user_email=user_email,
            filename=file.filename,
            filepath=filepath,
            size=size,
        )
        db.add(doc)
        db.commit()
        db.refresh(doc)
    except SQLAlchemyError as e:
        db.rollback()
        logger.error(f"Error making change to database: {e}")
        raise HTTPException(status_code=500, detail="Database error")
        

    try:
        chunks = parse_pdf(pdf_bytes)
        init_db(chunks, f"{user_email}_{file.filename}", OPENAI_API_KEY)
    except Exception as e:
        logger.error("Failed to build FAISS index for user %s, file %s: %s", user_email, file.filename, e)
        raise HTTPException(status_code=500, detail="An unexpected error occurred")

    return {
        "id": doc.id,
        "filename": doc.filename,
        "size": doc.size,
        "uploaded_at": doc.uploaded_at.isoformat(),
    }




@router.get("/api/documents")
def list_documents(db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):

    user_email = current_user["username"]

    docs = None 
    try:
        docs = db.query(Document).filter(Document.user_email == user_email).all()
    except SQLAlchemyError as e:
        logger.error(f"Error querying user: {e}")
        raise HTTPException(status_code=500, detail="Database Error")
    
    return [
        {
            "id": doc.id,
            "filename": doc.filename,
            "size": doc.size,
            "uploaded_at": doc.uploaded_at.isoformat(),
        }
        for doc in docs
    ]


@router.get("/api/documents/{doc_id}/export")
def export_document(doc_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):


    user_email = current_user["username"]

    doc = None
    try: 
        doc = db.query(Document).filter(Document.id == doc_id, Document.user_email == user_email).first()
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
    except SQLAlchemyError as e:
        logger.error(f"Error querying the database: {e}")
        raise HTTPException(status_code=500, detail="Database Error")

    try:
        with open(doc.filepath, "rb") as f:
            pdf_bytes = f.read()
        chunks = parse_pdf(pdf_bytes)
    except Exception as e:
        logger.error("Failed to read/parse PDF for doc %s: %s", doc_id, e)
        raise HTTPException(status_code=500, detail="An unexpected error occurred")

    
    semesters = {}
    for _, metadata in chunks:
        if metadata["type"] == "semester":
            semesters[metadata["semester"]] = {
                "term_gpa": metadata["term_gpa"],
                "cum_gpa": metadata["cum_gpa"],
                "courses": [],
            }
        elif metadata["type"] == "course":
            label = metadata["semester"]
            if label in semesters:
                semesters[label]["courses"].append(metadata)

    wb = Workbook()
    ws = wb.active
    ws.title = "Transcript"

 
    sem_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    gpa_font = Font(bold=True, italic=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    summary_fill = PatternFill("solid", fgColor="E2EFDA")
    center = Alignment(horizontal="center")
    thin = Side(style="thin")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_borders(start_row, end_row, start_col, end_col):
        for r in range(start_row, end_row + 1):
            for c in range(start_col, end_col + 1):
                ws.cell(row=r, column=c).border = cell_border

    COURSE_HEADERS = ["Course", "Description", "Grade", "Credits Attempted", "Credits Earned", "Points"]

    row = 1

    for sem_label, sem_data in semesters.items():
        sem_start = row

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COURSE_HEADERS))
        cell = ws.cell(row=row, column=1, value=sem_label)
        cell.font = sem_font
        cell.fill = PatternFill("solid", fgColor="BDD7EE")
        cell.alignment = center
        row += 1

        for col, h in enumerate(COURSE_HEADERS, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        row += 1

        for course in sem_data["courses"]:
            ws.cell(row=row, column=1, value=course["course"])
            ws.cell(row=row, column=2, value=course["description"])
            ws.cell(row=row, column=3, value=course["grade"]).alignment = center
            ws.cell(row=row, column=4, value=course["attempted"]).alignment = center
            ws.cell(row=row, column=5, value=course["earned"]).alignment = center
            ws.cell(row=row, column=6, value=course["points"]).alignment = center
            row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        gpa_cell = ws.cell(row=row, column=1, value=f"Term GPA: {sem_data['term_gpa']}   |   Cumulative GPA: {sem_data['cum_gpa']}")
        gpa_cell.font = gpa_font
        gpa_cell.alignment = center

        apply_borders(sem_start, row, 1, len(COURSE_HEADERS))
        row += 2  


    summary_start = row

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    summary_title = ws.cell(row=row, column=1, value="GPA Summary")
    summary_title.font = Font(bold=True, size=12)
    summary_title.fill = PatternFill("solid", fgColor="70AD47")
    summary_title.alignment = center
    row += 1

    for col, h in enumerate(["Semester", "Term GPA", "Cumulative GPA"], 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = header_font
        c.fill = summary_fill
        c.alignment = center
    row += 1

    for sem_label, sem_data in semesters.items():
        ws.cell(row=row, column=1, value=sem_label)
        ws.cell(row=row, column=2, value=sem_data["term_gpa"]).alignment = center
        ws.cell(row=row, column=3, value=sem_data["cum_gpa"]).alignment = center
        row += 1

    term_gpas = [float(d["term_gpa"]) for d in semesters.values() if d["term_gpa"]]
    avg_gpa = round(sum(term_gpas) / len(term_gpas), 3) if term_gpas else "N/A"
    ws.cell(row=row, column=1, value="Average").font = Font(bold=True)
    ws.cell(row=row, column=2, value=avg_gpa).alignment = center
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3, value="—").alignment = center

    apply_borders(summary_start, row, 1, 3)
    row += 1

    col_widths = [30, 40, 10, 20, 15, 10]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


    buffer = None 
    try: 
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
    except Exception as e:
        logger.error(f"Failed to export to workbook: {e}")
        raise HTTPException(status_code=500, detail="Something went wrong")

    export_filename = doc.filename.replace(".pdf", ".xlsx")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{export_filename}\""}
    )




@router.delete("/api/documents/{doc_id}")
def delete_document(doc_id: int, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):


    user_email = current_user["username"]

    doc = None
    try:
        doc = db.query(Document).filter(Document.id == doc_id, Document.user_email == user_email).first()
    except SQLAlchemyError as e:
        logger.error(f"Error fetching user: {e}")
        raise HTTPException(status_code=500, detail="Database Error")

    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    if os.path.exists(doc.filepath):
        os.remove(doc.filepath)

    delete_index(f"{user_email}_{doc.filename}")

    try:
        db.query(Message).filter(Message.document_id == doc_id).delete()
        db.delete(doc)
        db.commit()
    except SQLAlchemyError as e:
        db.rollback()
        logger.error(f"Error deleting document: {e}")
        raise HTTPException(status_code=500, detail="Database Error")

    return {"message": "Document deleted"}
